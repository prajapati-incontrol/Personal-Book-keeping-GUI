import tkinter as tk
from tkinter import ttk # a package to get the themed widgets of tkinter
import openpyxl
from tkcalendar import DateEntry
from datetime import datetime
import matplotlib.pyplot as plt


class MyGUI():
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Book-Keeper")
        self.combo_list = ["Date", "Payee", "Description","Price", "A/C"]
        self.style = ttk.Style(self.root)
        self.root.tk.call("source", "forest-light.tcl")
        self.root.tk.call("source", "forest-dark.tcl")
        self.style.theme_use("forest-dark")
        
        self.frame = ttk.Frame(self.root)
        self.frame.pack()
        
        ########### status frame #############
        self.status_frame = ttk.LabelFrame(self.frame, text = "All A/C status")
        self.status_frame.grid(row = 0, column = 0, padx = 20, pady = 10)
        
        # ABN Account
        self.abn_amro = ttk.Label(self.status_frame, text = "Your Bank A/C")
        self.abn_amro.grid(row = 0, column = 0, sticky = "ew", padx = 10, pady = 10)
        # load the account balance
        path = r"path\to\your\latest_pocket.xlsx"
        self.balance_abn = openpyxl.load_workbook(path)
        # first sheet by default
        self.sheet = self.balance_abn.active
        # read values
        list_values = list(self.sheet.values)
        filtered_values = [tuple(value for value in row)[:6] 
                           for row in list_values]
        self.ac_amt = tk.Label(self.status_frame, text = "€ " + str(filtered_values[-1][5]))
        self.ac_amt.grid(row = 0, column = 1, sticky = "ew")
        
        ############# Entry Frame ##############
        self.entry_frame = ttk.LabelFrame(self.frame, text = "Insert Entry")
        self.entry_frame.grid(row = 1, column = 0, padx = 10, pady = 10)
        
        # entry 1: date
        self.date_label = ttk.Label(self.entry_frame, text = "Date")
        self.date_label.grid(row = 0, column = 0, padx = 10, pady = 10)
        self.date_entry = DateEntry(self.entry_frame, 
                                    width=12,
                                    background='darkblue',
                                    foreground='white', 
                                    borderwidth=2, 
                                    date_pattern = 'dd/mm/y')
        
        self.date_entry.grid(row = 0, column = 1, sticky = "ew", padx = 10, pady = 10)
        
        # entry 2: payee
        combo_list = ["Lidl", "Action", 
                      "Albert Heijn", "Jumbo", 
                      "NS","Lebara",
                      "Amazon", "Decathlon", 
                      "Nisarga Fresh","Bahar","Miscellaneous"]
        self.payee_label = ttk.Label(self.entry_frame, text = "Payee")
        self.payee_label.grid(row = 1, column = 0, padx = 10, pady =10)
        self.payee = ttk.Combobox(self.entry_frame, values = combo_list)
        self.payee.current(0)
        self.payee.grid(row = 1, column = 1, sticky = "ew", padx = 10, pady = 10)
        
        # Bind the selection event << because it is a virtual event
        self.payee.bind("<<ComboboxSelected>>", self.on_payee_select)
        
        self.separator = ttk.Separator(self.entry_frame)
        self.separator.grid(row = 3, column = 0, columnspan=3, padx = (10,10), pady = 10, sticky = 'ew')
        
        # entry 3: Item Name
        self.item_entry = ttk.Entry(self.entry_frame)
        self.item_entry.grid(row = 4, column = 0, sticky = "ew", padx = 10, pady = 10)
        self.item_entry.insert(0,"Item Name")
        self.item_entry.bind("<FocusIn>", lambda e: self.item_entry.delete('0','end'))
        
        # entry 4: Item Price
        self.item_price = ttk.Entry(self.entry_frame)
        self.item_price.grid(row = 4, column = 1, sticky = "ew", padx = 10, pady = 10)
        self.item_price.insert(0,"Item Price")
        self.item_price.bind("<FocusIn>", lambda e: self.item_price.delete('0','end'))
        
        # create button: Insert Row
        self.button1 = ttk.Button(self.entry_frame, text = "Insert Row", command = self.insert_row)
        self.button1.grid(row = 4, column = 2, padx = 10, pady = 10, sticky = "ew")
        
        self.separator2 = ttk.Separator(self.entry_frame)
        self.separator2.grid(row = 5, column = 0, columnspan=3, padx = (10,10), pady = 10, sticky = 'ew')
              
        
        # entry 5: method of payment
        self.method_label = ttk.Label(self.entry_frame, text = "Method of Payment")
        self.method_label.grid(row = 6, column = 0, padx = 10, pady = 10, sticky = "ew")
        
        method_list = ["Revolut", "Debit Card", "Cash","Tikkie"]
        self.method = ttk.Combobox(self.entry_frame, values = method_list)
        self.method.current(0)
        self.method.grid(row = 6, column = 1, padx = 10, pady = 10, sticky = "ew")
        
        
        ############## tree frame ##############
        self.treeframe = ttk.Frame(self.frame)
        self.treeframe.grid(row = 0,rowspan = 2, column = 1, pady = 10)
        self.treescroll = ttk.Scrollbar(self.treeframe)
        self.treescroll.pack(side = "right", fill = "y")
        
        self.treecols = ["Date", "Payee", "Item","Price","Method","Balance"]
        self.treeview = ttk.Treeview(self.treeframe, show = "headings", 
                                     yscrollcommand=self.treescroll.set, 
                                     columns=self.treecols,
                                     height = 15)
        
        self.treeview.column("Date", width = 150)
        self.treeview.column("Payee", width = 100)
        self.treeview.column("Item", width = 100)
        self.treeview.column("Price", width = 100)
        self.treeview.column("Method", width = 100)
        self.treeview.column("Balance", width = 100)
        
        self.treeview.pack()
        self.treescroll.config(command = self.treeview.yview)
        
        self.load_data()
        
        self.root.mainloop()
        
    def insert_row(self):
               
        # 1. take the data from gui
        #print(type(self.date_entry.get()))
        new_date = datetime.strptime(self.date_entry.get(), '%d/%m/%Y').date()
        #print(new_date)
        new_payee = self.payee.get()
        new_item = self.item_entry.get()
        new_value = float(self.item_price.get())
        new_method = self.method.get()
        #print(float(self.current_balance)), print(float(new_value))
        self.current_balance = float(self.current_balance) - float(new_value)
        #print(self.current_balance)
                
        # print(new_date, new_payee, 
        #       new_item, new_value, 
        #       new_method, self.current_balance)
        
        # check if the month changes
        # Parse the date entry
        month_name = new_date.strftime('%B')
        #print(month_name)
        self.latest_book = openpyxl.load_workbook(self.path)
        # Check if a sheet for this month exists
        if month_name not in self.latest_book.sheetnames:
            # Create new sheet for that month
            self.latest_book.create_sheet(title=month_name)
            print(f"Created new sheet: {month_name}")
            # get new sheet 
            self.latest_sheet = self.latest_book[month_name]
            
            # setup the headings in new sheet
            self.latest_sheet.append(self.treecols)
            
            # 2. insert row in excel sheet
            row_values = [new_date, new_payee,
                          new_item, new_value, 
                          new_method, self.current_balance]
            self.latest_sheet.append(row_values)
            self.latest_book.save(self.path)
            
            # empty the tree view
            for item in self.treeview.get_children():
                self.treeview.delete(item)
            # 3 insert row in the tree view
            self.treeview.insert('',tk.END, values = row_values)
                
        else: # no need to make new sheet 
            self.latest_sheet = self.latest_book[month_name]
            row_values = [new_date, new_payee,
                          new_item, new_value, 
                          new_method, self.current_balance]
            
            self.latest_sheet.append(row_values)
            self.latest_book.save(self.path)
            
            # 3 insert row in the tree view
            self.treeview.insert('',tk.END, values = row_values)
            
        # 4 reset tabs
        self.item_entry.delete(0,"end")
        self.item_entry.insert(0,"Item Name")
        self.item_price.delete(0,"end")
        self.item_price.insert(0,"Item Price")
        self.ac_amt = tk.Label(self.status_frame, text = "€ " + str(self.current_balance))
        self.ac_amt.grid(row = 0, column = 1, sticky = "ew")
        
    
    def load_data(self):
        self.path = r"path\to\your\latest_pocket.xlsx"
        workbook = openpyxl.load_workbook(self.path)
        # first sheet by default
        sheet = workbook.active
        
        # read values
        list_values = list(sheet.values)
        filtered_values = [tuple(value for value in row)[:6] 
                           for row in list_values]
        # print(filtered_values)
        for col_name in filtered_values[0]:
            self.treeview.heading(col_name, text = col_name)

        # print(filtered_values)
        self.current_balance = filtered_values[-1][-1]
        # print(self.initial_balance)
        
        # Iteratve over the filtered_values starting from index 1
        for i in range(1, len(filtered_values)):
            temp_list = list(filtered_values[i])
            temp_list[0] = temp_list[0].date()
            filtered_values[i] = tuple(temp_list)
            
        
        for value_tuple in filtered_values[1:]:
            #print(value_tuple)
            self.treeview.insert('', tk.END, values = value_tuple)
        
    def on_payee_select(self, event):
        if self.payee.get() == "Miscellaneous":
            self.misc_label = ttk.Label(self.entry_frame, text = "Please specify:")
            self.misc_label.grid(row = 2, column = 0, padx = 10, pady = 10)
            
            self.misc_entry = ttk.Entry(self.entry_frame)
            self.misc_entry.grid(row = 2, column = 1, padx = 10, pady = 10, sticky = "ew")
        else:
            if hasattr(self, "misc_label"):
                self.misc_label.grid_forget()
                self.misc_entry.grid_forget()
        
MyGUI()